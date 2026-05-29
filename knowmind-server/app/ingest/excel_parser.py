from __future__ import annotations

import csv
import io
from pathlib import Path

from app.ingest.types import PageText, ParseResult


def _infer_column_type(values: list[str]) -> str:
    non_empty = [v for v in values if v.strip()]
    if not non_empty:
        return "空"
    nums = 0
    for v in non_empty:
        try:
            float(v.replace(",", ""))
            nums += 1
        except ValueError:
            pass
    if nums == len(non_empty):
        return "数值"
    if all(len(v) <= 20 for v in non_empty[:20]):
        return "分类/短文本"
    return "文本"


def _sheet_to_markdown(sheet_name: str, rows: list[list[str]]) -> str:
    if not rows:
        return ""
    col_count = max(len(r) for r in rows)
    normalized = [r + [""] * (col_count - len(r)) for r in rows]
    headers = normalized[0]
    data_rows = normalized[1:] if len(normalized) > 1 else []

    lines: list[str] = [f"## 工作表：{sheet_name}", ""]
    # 数据字典
    lines.append("### 字段说明")
    lines.append("")
    lines.append("| 字段 | 类型推断 | 示例值 | 非空数 |")
    lines.append("| --- | --- | --- | --- |")
    for ci, header in enumerate(headers):
        col_vals = [r[ci] if ci < len(r) else "" for r in data_rows]
        non_empty = [v for v in col_vals if v.strip()]
        sample = non_empty[0][:40] if non_empty else "—"
        col_type = _infer_column_type(col_vals)
        lines.append(f"| {header or f'列{ci+1}'} | {col_type} | {sample} | {len(non_empty)} |")

    lines.extend(["", "### 数据概要", ""])
    lines.append(f"- 行数（含表头）：{len(normalized)}")
    lines.append(f"- 列数：{col_count}")
    if data_rows:
        lines.append(f"- 数据行：{len(data_rows)}")

    lines.extend(["", "### 数据预览", ""])
    preview_rows = normalized[: min(11, len(normalized))]
    for i, row in enumerate(preview_rows):
        line = "| " + " | ".join(c.replace("|", "\\|") for c in row) + " |"
        lines.append(line)
        if i == 0:
            lines.append("| " + " | ".join("---" for _ in row) + " |")
    if len(normalized) > 11:
        lines.append(f"\n*（共 {len(normalized)} 行，仅展示前 10 行数据）*")

    return "\n".join(lines)


def parse_csv(path: str, filename: str | None = None) -> ParseResult:
    raw = Path(path).read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    if text is None:
        raise ValueError("无法识别 CSV 编码")

    reader = csv.reader(io.StringIO(text))
    rows = [[c.strip() for c in row] for row in reader if any(c.strip() for c in row)]
    base = (filename or Path(path).name).rsplit(".", 1)[0]
    content = _sheet_to_markdown(base, rows)
    pages = [PageText(page_index=0, text=content)]
    summary = f"CSV 表格 {len(rows)} 行 × {max((len(r) for r in rows), default=0)} 列"
    return ParseResult(pages=pages, title=base[:200], summary=summary, content=content)


def parse_excel(path: str, filename: str | None = None) -> ParseResult:
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(path, read_only=True, data_only=True)
    sections: list[str] = []
    pages: list[PageText] = []
    sheet_count = len(wb.sheetnames)
    try:
        for si, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in row]
                if any(cells):
                    rows.append(cells)
            md = _sheet_to_markdown(sheet_name, rows)
            if md:
                sections.append(md)
                pages.append(PageText(page_index=si, text=md))
    finally:
        wb.close()

    content = "\n\n---\n\n".join(sections)
    base = (filename or Path(path).name).rsplit(".", 1)[0]
    summary = f"Excel {sheet_count} 个工作表" if sections else None
    return ParseResult(pages=pages or [PageText(page_index=0, text="")], title=base[:200], summary=summary, content=content)


def parse_xls(path: str, filename: str | None = None) -> ParseResult:
    try:
        import xlrd  # noqa: PLC0415
    except ImportError as e:
        raise ValueError("不支持旧版 .xls，请另存为 .xlsx 后上传") from e

    book = xlrd.open_workbook(path)
    sections: list[str] = []
    pages: list[PageText] = []
    for si in range(book.nsheets):
        sheet = book.sheet_by_index(si)
        rows: list[list[str]] = []
        for ri in range(sheet.nrows):
            cells = [str(sheet.cell_value(ri, ci)).strip() for ci in range(sheet.ncols)]
            if any(cells):
                rows.append(cells)
        md = _sheet_to_markdown(sheet.name, rows)
        if md:
            sections.append(md)
            pages.append(PageText(page_index=si, text=md))
    content = "\n\n---\n\n".join(sections)
    base = (filename or Path(path).name).rsplit(".", 1)[0]
    return ParseResult(pages=pages, title=base[:200], summary=f"Excel {book.nsheets} 个工作表", content=content)
